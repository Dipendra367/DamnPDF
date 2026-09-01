import asyncio
import uuid
import hashlib
from pathlib import Path

import pdfplumber
import pandas as pd
from fastapi import APIRouter, UploadFile, File, HTTPException, Query
from fastapi.responses import FileResponse
from openpyxl.drawing.image import Image as XLImage
from openpyxl import load_workbook

from core.office_convert import TEMP_DIR
from core.config import MAX_FILE_SIZE
from core.limits import conversion_semaphore

router = APIRouter()

IMAGE_COLUMN = "J"  # images always placed here, well clear of text columns


def _table_signature(table) -> str:
    return hashlib.md5(str(table).encode("utf-8", errors="ignore")).hexdigest()


def build_workbook(input_path: Path, output_path: Path, img_dir: Path, layout: str):
    """Blocking worker: parses the PDF and writes the xlsx."""
    sheets = []          # list of (sheet_name, df_or_None)
    sheet_images = {}    # sheet_name -> list of image file paths
    seen_table_hashes = set()  # dedup across the whole document

    with pdfplumber.open(input_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            tables = page.extract_tables()
            page_added_content = False

            if tables:
                for t_idx, table in enumerate(tables, start=1):
                    if not table or len(table) <= 1:
                        continue

                    sig = _table_signature(table)
                    if sig in seen_table_hashes:
                        continue  # skip duplicate table (pdfplumber double-detect quirk)
                    seen_table_hashes.add(sig)

                    df = pd.DataFrame(table[1:], columns=table[0])
                    sheet_name = f"Page{page_num}_Table{t_idx}"[:31]
                    sheets.append((sheet_name, df))
                    page_added_content = True
            else:
                text = page.extract_text()
                if text:
                    sig = hashlib.md5(text.encode("utf-8", errors="ignore")).hexdigest()
                    if sig not in seen_table_hashes:
                        seen_table_hashes.add(sig)
                        lines = [line for line in text.split("\n") if line.strip()]
                        df = pd.DataFrame(lines, columns=["Content"])
                        sheet_name = f"Page{page_num}_Text"[:31]
                        sheets.append((sheet_name, df))
                        page_added_content = True

            # Extract images/figures on this page
            if page.images:
                img_sheet_name = sheets[-1][0] if page_added_content else f"Page{page_num}_Figures"[:31]
                saved_paths = []
                for i_idx, img in enumerate(page.images, start=1):
                    try:
                        bbox = (img["x0"], img["top"], img["x1"], img["bottom"])
                        cropped = page.crop(bbox).to_image(resolution=150)
                        img_path = img_dir / f"p{page_num}_img{i_idx}.png"
                        cropped.save(str(img_path))
                        saved_paths.append(img_path)
                    except Exception:
                        continue

                if saved_paths:
                    if not page_added_content:
                        sheets.append((img_sheet_name, None))
                    sheet_images.setdefault(img_sheet_name, []).extend(saved_paths)

    if not sheets:
        raise HTTPException(
            status_code=422,
            detail="No extractable content found in the PDF.",
        )

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        if layout == "one":
            start_row = 0
            for sheet_name, df in sheets:
                header_df = pd.DataFrame([[sheet_name]])
                header_df.to_excel(writer, sheet_name="Sheet1", index=False, header=False, startrow=start_row)
                start_row += 2
                if df is not None:
                    df.to_excel(writer, sheet_name="Sheet1", index=False, startrow=start_row)
                    n_images = len(sheet_images.get(sheet_name, []))
                    # gap large enough for either the text rows or the images stacked beside them
                    start_row += max(len(df), n_images * 22) + 4
                else:
                    start_row += 2
        else:
            for sheet_name, df in sheets:
                if df is not None:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                else:
                    writer.book.create_sheet(sheet_name)

    # Second pass: embed images, always in a dedicated column so they never overlap text
    wb = load_workbook(output_path)

    if layout == "one":
        ws = wb["Sheet1"]
        # Re-walk the blocks to know where each sheet's block started, so images
        # land next to their own text block instead of at the bottom.
        start_row = 0
        for sheet_name, df in sheets:
            start_row += 2
            block_start = start_row
            if df is not None:
                n_images = len(sheet_images.get(sheet_name, []))
                row_cursor = block_start
                for img_path in sheet_images.get(sheet_name, []):
                    xl_img = XLImage(str(img_path))
                    xl_img.width, xl_img.height = min(xl_img.width, 350), min(xl_img.height, 350)
                    ws.add_image(xl_img, f"{IMAGE_COLUMN}{row_cursor}")
                    row_cursor += 22
                start_row += max(len(df), n_images * 22) + 4
            else:
                row_cursor = block_start
                for img_path in sheet_images.get(sheet_name, []):
                    xl_img = XLImage(str(img_path))
                    xl_img.width, xl_img.height = min(xl_img.width, 350), min(xl_img.height, 350)
                    ws.add_image(xl_img, f"{IMAGE_COLUMN}{row_cursor}")
                    row_cursor += 22
                start_row += 2
    else:
        for sheet_name, img_paths in sheet_images.items():
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)
            row_cursor = 1
            for img_path in img_paths:
                xl_img = XLImage(str(img_path))
                xl_img.width, xl_img.height = min(xl_img.width, 350), min(xl_img.height, 350)
                ws.add_image(xl_img, f"{IMAGE_COLUMN}{row_cursor}")
                row_cursor += 22

    wb.save(output_path)


@router.post("/pdf-to-excel")
async def pdf_to_excel(
    file: UploadFile = File(...),
    layout: str = Query("multiple", enum=["one", "multiple"]),
):
    if file.content_type != "application/pdf":
        raise HTTPException(status_code=400, detail="File must be a PDF")

    content = await file.read()

    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=413, detail="File too large. Max size is 100MB.")

    job_id = uuid.uuid4().hex[:8]
    input_path = TEMP_DIR / f"{job_id}_input.pdf"
    output_path = TEMP_DIR / f"{job_id}_output.xlsx"
    img_dir = TEMP_DIR / f"{job_id}_imgs"
    img_dir.mkdir(exist_ok=True)
    input_path.write_bytes(content)

    try:
        async with conversion_semaphore:
            await asyncio.to_thread(build_workbook, input_path, output_path, img_dir, layout)

        return FileResponse(
            path=output_path,
            filename="converted.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=500, detail="PDF to Excel conversion failed")
    finally:
        input_path.unlink(missing_ok=True)
        for f in img_dir.glob("*"):
            f.unlink(missing_ok=True)
        img_dir.rmdir()